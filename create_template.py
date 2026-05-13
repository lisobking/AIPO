from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

def create_po_template(template_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase_Order"

    # 1. 제목 및 로고 영역 (디자이너 감각 반영)
    ws.merge_cells('A1:E2')
    ws['A1'] = "발 주 서 (PURCHASE ORDER)"
    ws['A1'].font = Font(size=20, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A1'].fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    # 2. 기본 정보 헤더
    headers = ["No", "내부 품목코드", "업체 품목명", "수량", "단가", "합계"]
    ws.append(headers)
    for cell in ws[3]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # 3. 테두리 설정 (디테일)
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # 샘플 데이터 영역 서식 미리 지정 (100행까지)
    for row in range(3, 100):
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border

    wb.save(template_path)
    print(f"✅ 디자이너: 표준 발주서 템플릿 생성 완료. ({template_path.name})")

if __name__ == "__main__":
    TEMPLATE_PATH = Path(__file__).parent / "template" / "Standard_PO_Template.xlsx"
    TEMPLATE_PATH.parent.mkdir(exist_ok=True)
    create_po_template(TEMPLATE_PATH)
