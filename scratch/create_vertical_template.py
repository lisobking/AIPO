import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from pathlib import Path

def create_vertical_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vertical Purchase Order"
    
    # 1. 스타일 정의
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    base_font = Font(name='맑은 고딕', size=10)

    # 2. 제목 및 헤더 (가로형과 통일성 유지)
    ws.merge_cells('B2:G3')
    ws['B2'] = "발      주      서 (Vertical)"
    ws['B2'].font = Font(name='맑은 고딕', size=24, bold=True)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # 수신, 참조 영역 (C10, C11)
    ws['B10'] = "수 신 :"
    ws['B11'] = "참 조 :"
    ws.merge_cells('C10:E10')
    ws.merge_cells('C11:E11')

    # 3. 본문 헤더 (17행)
    headers = ["NO", "분 류", "품 명 및 규 격", "수 량", "단 가", "합 계"]
    cols = ['B', 'C', 'D', 'E', 'F', 'G']
    for col, text in zip(cols, headers):
        cell = ws[f"{col}17"]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 4. 세로형 특화 데이터 영역 (18-50행)
    for r in range(18, 51):
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = base_font
            # 세로형의 핵심: 모든 셀에 자동 줄바꿈 및 상단 정렬 적용
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left' if c==4 else 'center')

    # 5. 너비 조정 (세로형은 품명 칸을 더 넓게)
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 55 # 품명 칸 대폭 확장
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15

    # 템플릿 저장
    template_path = Path("template/Vertical_PO_Template.xlsx")
    template_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(template_path)
    print(f"✅ 세로형 특화 템플릿 생성 완료: {template_path}")

if __name__ == "__main__":
    create_vertical_template()
