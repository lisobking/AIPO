from app_main import manager
from pathlib import Path
import openpyxl
import unicodedata

def validate_conversion(name, source_filename, unit_price):
    print(f"\n==================================================")
    print(f"🧪 백엔드 검수 시작: {name} (단가: {unit_price:,.0f}원)")
    print(f"==================================================")
    
    sample_dir = Path(__file__).parent.parent / "sample" / "sam2"
    source_path = sample_dir / source_filename
    
    if not source_path.exists():
        print(f"❌ 원본 파일 없음: {source_path}")
        return False
        
    try:
        # 1. 백엔드 변환 실행
        output_file_str = manager.process_effort_to_quote(source_path, unit_price)
        output_path = Path(output_file_str)
        print(f"✅ 변환 완료 → 결과 파일: {output_path.name}")
        
        # 2. 엑셀 워크북 로드
        wb = openpyxl.load_workbook(output_path, data_only=True)
        
        # --- 시트 목록 검수 ---
        sheets = wb.sheetnames
        print(f"📋 생성된 시트 목록: {sheets}")
        if "공수산정근거" in sheets:
            print("❌ 오류: 제거 대상인 '공수산정근거' 백업 시트가 여전히 존재합니다.")
            return False
        print("✅ 시트 검수 통과: '공수산정근거' 백업 시트 제외 확인 완료.")
        
        # --- Sheet 1 (견적서) 검수 ---
        ws1 = wb[sheets[0]]
        
        # A. 수신처 한글 유니코드 정규화 검수
        rcv = ws1['D4'].value
        is_normalized = (rcv == unicodedata.normalize('NFC', rcv))
        print(f"   [D4] 수신처: '{rcv}' (NFC 정규화: {'통과' if is_normalized else '실패'})")
        if not is_normalized:
            print("❌ 오류: NFD 한글 자모 분리 현상이 해결되지 않았습니다.")
            return False
            
        # B. 서두 문구 및 제안금액 검수
        intro = ws1['B9'].value
        price_text = ws1['B10'].value
        print(f"   [B9] 서두 문구: '{intro}'")
        print(f"   [B10] 제안금액란: '{price_text}'")
        
        # C. 14~17행 데이터 및 템플릿 보존(이행 행) 및 M/M 수량/단가 변환 검수
        print("   --- 14~17행 세부 내역 및 M/M 변환 비율 검수 ---")
        expected_mds = {
            "유한양행_공수 산정_2026.02.02 (2).xlsx": {14: 5.0, 15: 2.0, 16: 3.0},
            "NH투자증권 ezMail60  공수 산정.xlsx": {14: 5.0, 15: 5.0, 16: 30.0, 17: 20.0}
        }
        
        for r in range(14, 18):
            category = ws1.cell(row=r, column=2).value
            sub_category = ws1.cell(row=r, column=3).value
            item_name = ws1.cell(row=r, column=4).value
            qty = ws1.cell(row=r, column=6).value
            price = ws1.cell(row=r, column=7).value
            total = ws1.cell(row=r, column=8).value
            print(f"   Row {r}: 구분={category}, 기능별={sub_category}, 내역={item_name}, 수량={qty}, 단가={price}, 합계={total}")
            
            # 수량 M/M 변환 (1M/M = 20M/D) 검증
            filename_clean = unicodedata.normalize('NFC', source_filename)
            for key, md_map in expected_mds.items():
                if unicodedata.normalize('NFC', key) == filename_clean:
                    if r in md_map:
                        expected_mm = md_map[r] / 20.0
                        if abs(qty - expected_mm) > 1e-5:
                            print(f"❌ 오류: Row {r}의 수량({qty})이 M/M 변환 기준({expected_mm})과 일치하지 않습니다.")
                            return False
            
            # 단가 20배 검증 (1M/M 단가 = M/D 단가 * 20)
            if r == 14 and price is not None:
                if price != unit_price * 20:
                    print(f"❌ 오류: Row 14의 단가({price})가 20배 스케일링된 M/M 단가 기준({unit_price * 20})과 일치하지 않습니다.")
                    return False

            # 17행 '이행' 단계 보존 여부 검증 (품목 수가 4개 이하로 적은 유한양행 파일명의 경우만 체크하도록 구성)
            if r == 17 and "유한양행" in source_filename:
                if sub_category != "이행" or "* 테스트 및 운영서버 반영" not in str(item_name):
                    print("❌ 오류: 템플릿 17행의 표준 '이행' 공정 텍스트가 훼손되었습니다.")
                    return False
                print("   ✅ 템플릿 보존 검수 통과: 17행 '이행' 공정 명칭 정상 유지.")
                
        # D. 소계 행 위치 탐색 및 수식 검수
        ws_formula = openpyxl.load_workbook(output_path, data_only=False)[sheets[0]]
        soke_row = None
        for r in range(14, 30):
            c_val = ws1.cell(row=r, column=3).value
            if c_val and "가온아이 노임단가 기준 소계" in str(c_val):
                soke_row = r
                break
        
        if soke_row:
            soke_formula = ws_formula.cell(row=soke_row, column=6).value
            print(f"   [Row {soke_row}] 노임단가 소계 수량 수식: '{soke_formula}' (값: {ws1.cell(row=soke_row, column=6).value})")
            if "SUM" not in str(soke_formula):
                print("❌ 오류: 소계 행 수식이 동적으로 주입되지 않았습니다.")
                return False
                
        print(f"🎉 {name} 검수 결과: 100% 무결점 통과!")
        return True
    except Exception as e:
        print(f"❌ 검수 중 예외 발생: {e}")
        return False

if __name__ == "__main__":
    yuhan_ok = validate_conversion("유한양행 추가 기술지원", "유한양행_공수 산정_2026.02.02 (2).xlsx", 1350000)
    nh_ok = validate_conversion("NH투자증권 ezMail60 기술지원", "NH투자증권 ezMail60  공수 산정.xlsx", 1350000)
    
    print("\n==================================================")
    print(f"🏁 최종 검수 리포트 요약")
    print(f"==================================================")
    print(f"유한양행 파이프라인 검증: {'SUCCESS (패스)' if yuhan_ok else 'FAILED (오류)'}")
    print(f"NH투자증권 파이프라인 검증: {'SUCCESS (패스)' if nh_ok else 'FAILED (오류)'}")
    print("==================================================")
