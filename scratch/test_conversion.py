from app_main import manager
from pathlib import Path
import openpyxl

sample_dir = Path(__file__).parent.parent / "sample" / "sam2"
yuhan_effort = sample_dir / "유한양행_공수 산정_2026.02.02 (2).xlsx"

print("--- Running Effort-to-Quote Pipeline for Yuhan ---")
output_file = manager.process_effort_to_quote(yuhan_effort, 1350000)
print(f"Generated Output Path: {output_file}")

# Let's inspect the generated file
print("\n--- Inspecting Generated Quotation Cells ---")
wb = openpyxl.load_workbook(output_file, data_only=True)
ws = wb.active

print(f"D4 ( 수신처): {ws['D4'].value}")
print(f"B9 ( 서두 문구): {ws['B9'].value}")
print(f"B10 (제안금액): {ws['B10'].value}")

print("\n--- Items Table (Rows 14 to 20) ---")
for r in range(13, 22):
    row_vals = [ws.cell(r, c).value for c in [2, 3, 4, 6, 7, 8, 9]]
    if any(v is not None for v in row_vals):
        print(f"Row {r:02d}: {row_vals}")

# Find Big_o row
bigo_row = None
for r in range(20, 35):
    b_val = ws.cell(row=r, column=2).value
    if b_val and '비' in str(b_val) and '고' in str(b_val):
        bigo_row = r
        break

if bigo_row:
    print(f"\n--- Remarks Table (Row {bigo_row}) ---")
    print(ws.cell(row=bigo_row, column=3).value)
