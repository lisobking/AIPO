import openpyxl
from pathlib import Path

def analyze_diff():
    er2_dir = Path("/Users/lisob/Desktop/project2/AutoPO_Project/error/er2")
    master_path = er2_dir / "유한양행 신규 양식 2종 포함 총 양식 6종 및 SSO 모듈 추가 기술지원_가온아이_2602107.xlsx"
    draft_path = Path("/Users/lisob/Desktop/project2/AutoPO_Project/workspace/04_output_quotes/견적초안_유한양행_공수 산정_2026.02.02 (2).xlsx")

    wb_master = openpyxl.load_workbook(master_path, data_only=False)
    wb_draft = openpyxl.load_workbook(draft_path, data_only=False)

    ws_master = wb_master.active
    ws_draft = wb_draft.active

    print("==================================================")
    print("🔍 엑셀 정밀 비교 엔진 (Excel Precision Diff Engine)")
    print("==================================================")
    print(f"🥇 마스터 파일 (완성되어야 할 본): {master_path.name}")
    print(f"🥈 현재 초안 파일 (완성된 본): {draft_path.name}\n")

    # 1. 시트 목록 비교
    print("📋 [1] 시트 목록 비교")
    print(f"   - 마스터 시트: {wb_master.sheetnames}")
    print(f"   - 초안 시트: {wb_draft.sheetnames}\n")

    # 2. 셀 병합 영역 비교
    print("🧩 [2] 셀 병합 비교 (Merged Cell Ranges)")
    master_merges = sorted([str(r) for r in ws_master.merged_cells.ranges])
    draft_merges = sorted([str(r) for r in ws_draft.merged_cells.ranges])
    
    only_in_master = [m for m in master_merges if m not in draft_merges]
    only_in_draft = [d for d in draft_merges if d not in master_merges]
    
    print(f"   - 마스터에만 있는 병합: {only_in_master}")
    print(f"   - 초안에만 있는 병합: {only_in_draft}\n")

    # 3. 셀 값 및 수식 1:1 디프 (40행 11열 범위)
    print("✏️ [3] 셀 값 및 수식 불일치 분석 (Formulas & Values)")
    diff_count = 0
    for r in range(1, 45):
        for c in range(1, 12):
            cell_ref = f"{openpyxl.utils.get_column_letter(c)}{r}"
            val_master = ws_master.cell(r, c).value
            val_draft = ws_draft.cell(r, c).value

            if val_master != val_draft:
                print(f"   📍 {cell_ref} -> 마스터: '{val_master}' | 초안: '{val_draft}'")
                diff_count += 1
    
    if diff_count == 0:
        print("   ✅ 모든 셀의 수식 및 값이 100% 완벽히 일치합니다!")
    else:
        print(f"   ⚠️ 총 {diff_count}개의 셀 값/수식 차이가 발견되었습니다.\n")

    # 4. 스타일(테두리, 색상, 정렬) 비교
    print("🎨 [4] 시각 서식 비교 (Borders, Colors, Alignment)")
    style_diffs = 0
    for r in range(1, 45):
        for c in range(1, 12):
            cell_ref = f"{openpyxl.utils.get_column_letter(c)}{r}"
            cell_m = ws_master.cell(r, c)
            cell_d = ws_draft.cell(r, c)

            # A. 정렬 비교
            align_m = cell_m.alignment.horizontal if cell_m.alignment else None
            align_d = cell_d.alignment.horizontal if cell_d.alignment else None
            if align_m != align_d:
                print(f"   📍 {cell_ref} [정렬] -> 마스터: '{align_m}' | 초안: '{align_d}'")
                style_diffs += 1

            # B. 테두리 비교 (특히 얇은 테두리와 아래 굵은 테두리 등)
            border_m = cell_m.border.bottom.style if cell_m.border and cell_m.border.bottom else None
            border_d = cell_d.border.bottom.style if cell_d.border and cell_d.border.bottom else None
            if border_m != border_d:
                print(f"   📍 {cell_ref} [하단 테두리] -> 마스터: '{border_m}' | 초안: '{border_d}'")
                style_diffs += 1

    if style_diffs == 0:
        print("   ✅ 모든 테두리 및 정렬 스타일이 일치합니다!")
    else:
        print(f"   ⚠️ 총 {style_diffs}개의 서식 스타일 차이가 발견되었습니다.\n")

if __name__ == "__main__":
    analyze_diff()
