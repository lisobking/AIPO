import openpyxl
from pathlib import Path

er2_dir = Path("/Users/lisob/Desktop/project2/AutoPO_Project/error/er2")
orig_path = er2_dir / "유한양행 신규 양식 2종 포함 총 양식 6종 및 SSO 모듈 추가 기술지원_가온아이_2602107.xlsx"
draft_path = er2_dir / "견적초안_유한양행_공수 산정_2026.02.02 (2).xlsx"

print("=== Analyzing Original Quotation Template ===")
wb_orig = openpyxl.load_workbook(orig_path, data_only=True)
print("Sheets in original:", wb_orig.sheetnames)
ws_orig = wb_orig.active
print(f"Active sheet: {ws_orig.title}")

# Print first 40 rows and first 10 columns to see where the data is
for r in range(1, 40):
    row_vals = [ws_orig.cell(r, c).value for c in range(1, 12)]
    # Filter out if all elements are None
    if any(v is not None for v in row_vals):
        print(f"Row {r:02d}: {row_vals}")

print("\n=== Analyzing Draft Quotation ===")
wb_draft = openpyxl.load_workbook(draft_path, data_only=True)
print("Sheets in draft:", wb_draft.sheetnames)
ws_draft = wb_draft.active
print(f"Active sheet: {ws_draft.title}")
for r in range(1, 25):
    row_vals = [ws_draft.cell(r, c).value for c in range(1, 8)]
    if any(v is not None for v in row_vals):
        print(f"Row {r:02d}: {row_vals}")
