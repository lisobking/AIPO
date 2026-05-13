import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from pathlib import Path

def create_po_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Order"
    
    # 1. 기본 스타일 정의
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    title_font = Font(name='맑은 고딕', size=24, bold=True)
    header_font = Font(name='맑은 고딕', size=11, bold=True)
    base_font = Font(name='맑은 고딕', size=10)

    # 2. 제목 영역
    ws.merge_cells('B2:G3')
    ws['B2'] = "발      주      서"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # 3. 메타데이터 영역 (상단 정보)
    # 발주번호, 발주일자 (D6, D7)
    ws['F6'] = "발주번호 :"
    ws['F7'] = "발주일자 :"
    ws['F6'].font = header_font
    ws['F7'].font = header_font
    
    # 수신, 참조 영역 (B10, B11)
    ws['B10'] = "수 신 :"
    ws['B11'] = "참 조 :"
    ws['B10'].font = header_font
    ws['B11'].font = header_font
    ws.merge_cells('C10:E10')
    ws.merge_cells('C11:E11')

    # 4. 본문 테이블 헤더 (17행)
    headers = ["NO", "분 류", "품 명 및 규 격", "수 량", "단 가", "합 계"]
    cols = ['B', 'C', 'D', 'E', 'F', 'G']
    
    for col, text in zip(cols, headers):
        cell = ws[f"{col}17"]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 5. 데이터 영역 가이드라인 (18-35행)
    for r in range(18, 36):
        for c in range(2, 8):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.font = base_font

    # 6. 컬럼 너비 조정
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 18

    # 템플릿 저장
    template_path = Path("template/Standard_PO_Template.xlsx")
    template_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(template_path)
    print(f"✅ 신규 표준 템플릿 생성 완료: {template_path}")

if __name__ == "__main__":
    create_po_template()
