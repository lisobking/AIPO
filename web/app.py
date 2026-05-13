import os
import shutil
from pathlib import Path
from flask import Flask, render_template, request, send_from_directory, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import pdfplumber
import re
from copy import copy

# 프로젝트 루트 및 경로 설정
PROJECT_ROOT = Path(__file__).parent.parent.resolve()
UPLOAD_FOLDER = PROJECT_ROOT / "workspace" / "01_input_quotes"
RESULT_FOLDER = PROJECT_ROOT / "workspace" / "02_output_pos"
# 새로 지정된 샘플 발주서를 표준 템플릿으로 사용 (자사 정보 보존용)
STD_TEMPLATE = PROJECT_ROOT / "template" / "Standard_PO_Template.xlsx"

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = str(UPLOAD_FOLDER)

# 디렉토리 보장
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)
RESULT_FOLDER.mkdir(parents=True, exist_ok=True)

class PDFQuoteProcessor:
    """개발자 2: 지능형 PDF 문맥 분석 엔진"""
    def process(self, pdf_path):
        with pdfplumber.open(pdf_path) as pdf:
            tables = [page.extract_table() for page in pdf.pages]
            
        items = []
        for table in tables:
            if not table: continue
            for row in table:
                if not row or not any(row): continue
                
                # 행 전체에서 가장 이름다운 텍스트 찾기
                name = ""
                price = 0
                noise = ["발신자", "주소", "종목", "사업부", "유효기간", "고객사", "견적명", "합계", "총액", "이건창호", "가온아이", "ITEM", "MODEL"]
                
                for val in row:
                    v_str = str(val).strip()
                    if len(v_str) > 2 and not any(k in v_str for k in noise):
                        if not name or len(v_str) > len(name): # 가장 긴 텍스트를 이름 후보로
                            name = v_str
                    
                    # 가격 탐색
                    try:
                        s = v_str.replace(',', '').replace('원', '')
                        if s.replace('.','').isdigit() and float(s) > 1000:
                            price = float(s)
                    except: pass

                if not name or any(k in name.upper() for k in ["TEL", "FAX", "EMAIL", "■", "★"]): continue
                
                details = []
                for val in row:
                    v_str = str(val).strip()
                    if '\n' in v_str or '-' in v_str or '기간' in v_str:
                        details.extend([d.strip() for d in v_str.split('\n') if len(d.strip()) > 2 and d != name])

                if price > 0 or any(k in name.upper() for k in ["유지보수", "EDITOR", "에디터", "SERVER", "서버", "MAINTENANCE"]):
                    if not any(item['main'] == name for item in items):
                        items.append({'main': name, 'details': details, 'qty': 1, 'price': price})
        return items

class WebPOManager:
    """개발자 1: 엑셀 무결성 제어 및 시스템 통합"""
    def find_header_row(self, file_path):
        df_raw = pd.read_excel(file_path, header=None)
        for i, row in df_raw.iterrows():
            row_text = "".join([str(val).replace(" ", "") for val in row if not pd.isna(val)])
            if any(k in row_text for k in ['품목', '품명', 'ITEM', 'DESCRIPTION']): return i
        return 0

    def convert(self, file_path):
        ext = file_path.suffix.lower()
        items_to_write = []
        
        # 1. 데이터 추출
        if ext == '.pdf':
            pdf_engine = PDFQuoteProcessor()
            items_to_write = pdf_engine.process(file_path)
        else:
            header_idx = self.find_header_row(file_path)
            input_df = pd.read_excel(file_path, header=header_idx)
            
            # 컬럼 매핑 (주인공인 '품명'을 최우선 탐색)
            cols = [str(c).replace(" ", "").upper() for c in input_df.columns]
            def find_col(aliases, default_idx):
                for alias in aliases:
                    for i, col in enumerate(cols):
                        if alias in col: return input_df.columns[i]
                return input_df.columns[default_idx]

            target_col = find_col(['품명', 'ITEMNAME', 'DESCRIPTION', '품목'], 2)
            qty_col = find_col(['수량', 'QTY', 'QUANTITY'], 5)
            price_col = find_col(['단가', 'PRICE', 'UNITPRICE', '공급가', '공급단가', 'PRICE', 'COST'], 4)

            def safe_float(val):
                if pd.isna(val): return 0.0
                try:
                    # 모든 화폐 기호 및 콤마 제거
                    s = re.sub(r'[^0-9.]', '', str(val))
                    return float(s) if s else 0.0
                except: return 0.0

            # 3. 데이터 패턴 자동 보정 (중요: 키워드 매핑이 의심될 때 숫자가 많은 열을 수량/단가로 재지정)
            numeric_counts = input_df.select_dtypes(include=['number']).count()
            if len(numeric_counts) >= 2:
                # 숫자가 가장 많은 두 열을 단가와 수량 후보로 검토
                top_numeric_cols = numeric_counts.sort_values(ascending=False).index.tolist()
                # 단가는 보통 금액이 크므로, 평균값이 큰 쪽을 가격으로 추정하는 지능형 로직
                col1_vals = pd.to_numeric(input_df[top_numeric_cols[0]], errors='coerce').fillna(0)
                col2_vals = pd.to_numeric(input_df[top_numeric_cols[1]], errors='coerce').fillna(0)
                col1_avg = col1_vals.mean()
                col2_avg = col2_vals.mean()
                
                # 기존 매핑이 부실할 경우(평균 0) 보정
                current_price_avg = pd.to_numeric(input_df[price_col], errors='coerce').fillna(0).mean()
                current_qty_avg = pd.to_numeric(input_df[qty_col], errors='coerce').fillna(0).mean()

                if current_price_avg == 0:
                    price_col = top_numeric_cols[0] if col1_avg > col2_avg else top_numeric_cols[1]
                if current_qty_avg == 0:
                    qty_col = top_numeric_cols[1] if col1_avg > col2_avg else top_numeric_cols[0]

            processed = []
            noise_keywords = ["견적문의", "TEL", "FAX", "이메일", "■", "★", "TOTAL", "합계", "문의", "<", ">", "WWW.", "HTTP"]

            for _, row in input_df.iterrows():
                # 품명 열 데이터를 강제로 문자열로 변환하여 에러 방지
                name = str(row[target_col]).strip() if not pd.isna(row[target_col]) else ""
                name_upper = name.upper()
                if not name or name_upper in ["ITEM", "QTY", "PRICE", "MODELNO.", "품목", "NO", "DESCRIPTION"]: continue
                if any(k in name_upper for k in noise_keywords): continue
                
                qty = safe_float(row[qty_col])
                price = safe_float(row[price_col])

                if qty > 0 or price > 0:
                    processed.append({'main': name, 'details': [], 'qty': qty, 'price': price})
                elif processed and len(name) > 1:
                    processed[-1]['details'].append(name)
            items_to_write = processed

        # 2. 템플릿 기반 주입 (자사 정보 보존 전략)
        output_filename = f"PO_Draft_{file_path.stem}.xlsx"
        output_path = RESULT_FOLDER / output_filename
        shutil.copy(STD_TEMPLATE, output_path)
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 2. 템플릿 기반 주입 (자사 정보 보존 전략)
        output_filename = f"PO_Draft_{file_path.stem}.xlsx"
        output_path = RESULT_FOLDER / output_filename
        shutil.copy(STD_TEMPLATE, output_path)
        
        wb = load_workbook(output_path)
        ws = wb.active
        
        # 🛡️ 보안 및 무결성: 병합 셀 에러 방지를 위해 기존 데이터 삭제 방식을 '공백 주입'으로 변경
        start_row = 18
        for r in range(start_row, start_row + 40):
            for c in range(2, 9):
                try: ws.cell(row=r, column=c).value = None
                except: pass

        from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
        
        # 🛡️ 병합 셀 투과 주입 함수
        def safe_write(row, col, value):
            try:
                ws.cell(row=row, column=col, value=value)
            except: pass

        # 디자이너의 조색
        thin_gray = Side(border_style="thin", color="D3D3D3")
        thin_black = Side(border_style="thin", color="000000")
        no_border = Side(border_style=None)
        
        body_border = Border(left=thin_black, right=thin_black, top=thin_gray, bottom=thin_gray)
        remarks_border = Border(left=no_border, right=no_border, top=no_border, bottom=no_border)

        current_row = start_row
        item_no = 1
        total_sum = 0
        
        for item in items_to_write:
            # 메인 행 주입
            if item['qty'] > 0 or item['price'] > 0:
                safe_write(current_row, 2, item_no)
                item_no += 1
            
            # C=구분, D=명칭 매핑
            if len(str(item['main'])) < 6:
                safe_write(current_row, 3, str(item['main']))
            else:
                safe_write(current_row, 4, str(item['main']))
            
            if item['qty'] > 0: safe_write(current_row, 5, item['qty'])
            if item['price'] > 0: safe_write(current_row, 6, item['price'])
            
            line_total = item['qty'] * item['price']
            if line_total > 0:
                safe_write(current_row, 7, line_total)
                total_sum += line_total
            
            for c in range(2, 8):
                try: ws.cell(row=current_row, column=c).border = body_border
                except: pass
            current_row += 1
            
            for detail in item['details']:
                safe_write(current_row, 4, str(detail).strip().lstrip('-').strip())
                try: ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='left', indent=1)
                except: pass
                for c in range(2, 8):
                    try: ws.cell(row=current_row, column=c).border = body_border
                    except: pass
                current_row += 1

        # 3. 합계 영역 (Yellow Zone)
        if total_sum > 0:
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            bold_font = Font(bold=True)
            
            for label, val in [("합  계(VAT 별도)", total_sum), ("합  계(VAT 포함)", int(total_sum * 1.1))]:
                safe_write(current_row, 3, label)
                try: ws.cell(row=current_row, column=3).font = bold_font
                except: pass
                safe_write(current_row, 6, int(val / 12))
                safe_write(current_row, 7, val)
                
                for c in range(2, 8):
                    try:
                        ws.cell(row=current_row, column=c).fill = yellow_fill
                        ws.cell(row=current_row, column=c).border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
                    except: pass
                current_row += 1

        # 4. 비고란 영역
        current_row += 1
        remarks = [
            ("계약기간", f": 2026년 1월 1일 ~ 2026년 12월 30일 (1년)"),
            ("대금지급", ": 매월 말 세금계산서 발행/원청 수금 후 익월 15일 현금 지급"),
            ("납품장소", ": 고객사 지정위치"),
            ("담당자", ": 가온아이 CS사업부 박지혜 (iii511@kaoni.com 02-2140-5884)"),
            ("참  조", ": 첨부 견적서")
        ]
        for label, content in remarks:
            safe_write(current_row, 3, label)
            safe_write(current_row, 4, content)
            try: 
                ws.cell(row=current_row, column=3).font = bold_font
                for c in range(2, 8):
                    ws.cell(row=current_row, column=c).border = remarks_border
            except: pass
            current_row += 1

        wb.save(output_path)
        return output_filename

        wb.save(output_path)
        return output_filename

        wb.save(output_path)
        return output_filename

@app.route('/')
def index(): return render_template('index.html')

@app.route('/convert', methods=['POST'])
def convert_file():
    file = request.files['file']
    file_path = UPLOAD_FOLDER / file.filename
    file.save(str(file_path))
    manager = WebPOManager()
    res_name = manager.convert(file_path)
    return jsonify({'success': True, 'filename': res_name})

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(str(RESULT_FOLDER), filename, as_attachment=True)

# 🛡️ 비즈니스 보안: 휘발성 데이터 정책 (10분 후 자동 삭제)
@app.before_request
def cleanup_temp_files():
    import time
    now = time.time()
    for folder in [UPLOAD_FOLDER, RESULT_FOLDER]:
        for f in folder.glob("*"):
            if f.name == ".gitkeep": continue
            # 생성된 지 10분이 지난 파일은 즉시 삭제
            if now - f.stat().st_mtime > 600:
                try: f.unlink()
                except: pass

if __name__ == '__main__':
    # Render 등 클라우드 환경의 포트 대응
    port = int(os.environ.get("PORT", 5001))
    app.run(host='0.0.0.0', port=port)
