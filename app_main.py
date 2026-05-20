import os
import sys
import logging
import sqlite3
import shutil
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, request, jsonify, send_file

PROJECT_ROOT = Path(__file__).parent.resolve()
INPUT_DIR = PROJECT_ROOT / "workspace" / "01_input_quotes"
OUTPUT_DIR = PROJECT_ROOT / "workspace" / "02_output_pos"
TEMPLATE_FILE = PROJECT_ROOT / "template" / "Standard_PO_Template.xlsx"

# 견적서 초안 생성 파이프라인 경로
EFFORT_INPUT_DIR = PROJECT_ROOT / "workspace" / "03_input_effort"
QUOTE_OUTPUT_DIR = PROJECT_ROOT / "workspace" / "04_output_quotes"
LOG_DIR = PROJECT_ROOT / "logs"
WEB_DIR = PROJECT_ROOT / "web"

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s',
                    handlers=[logging.FileHandler(LOG_DIR / "system_security.log", encoding='utf-8'),
                              logging.StreamHandler(sys.stdout)])
logger = logging.getLogger("ParkDirector")

app = Flask(__name__, static_folder=str(WEB_DIR), static_url_path='/static')

class AutoPOManager:
    def __init__(self):
        self.ensure_directories()
        
    def ensure_directories(self):
        for path in [INPUT_DIR, OUTPUT_DIR, LOG_DIR, EFFORT_INPUT_DIR, QUOTE_OUTPUT_DIR]:
            path.mkdir(parents=True, exist_ok=True)

    def find_header_row(self, df):
        for i, row in df.iterrows():
            if any('품목' in str(val) or 'ITEM' in str(val) for val in row):
                return i
        return 0

    def process_file(self, file_path):
        output_path = OUTPUT_DIR / f"PO_{file_path.stem}.xlsx"
        try:
            df_raw = pd.read_excel(file_path, header=None)
            header_idx = self.find_header_row(df_raw)
            input_df = pd.read_excel(file_path, header=header_idx)
            
            shutil.copy(TEMPLATE_FILE, output_path)
            wb = load_workbook(output_path)
            ws = wb.active
            
            target_col = next((c for c in input_df.columns if '품목' in str(c) or 'ITEM' in str(c)), input_df.columns[0])
            price_col = next((c for c in input_df.columns if '단가' in str(c) or 'PRICE' in str(c)), None)
            
            start_row = 16 
            
            r_idx = 0
            for _, row in input_df.iterrows():
                item_name = row[target_col]
                if pd.isna(item_name) or str(item_name).strip() == "" or "ITEM" in str(item_name): continue
                price = row[price_col] if price_col else 0
                
                # 박부장: 동적 행 삽입
                if r_idx >= 10:
                    ws.insert_rows(start_row + r_idx)
                
                ws.cell(row=start_row + r_idx, column=2, value=r_idx + 1)
                ws.cell(row=start_row + r_idx, column=3, value=f"AUTO-CD-{r_idx}")
                ws.cell(row=start_row + r_idx, column=4, value=item_name)
                ws.cell(row=start_row + r_idx, column=5, value=1)
                ws.cell(row=start_row + r_idx, column=6, value=price)
                ws.cell(row=start_row + r_idx, column=7, value=price)
                r_idx += 1
                
            wb.save(output_path)
            return str(output_path)
        except Exception as e:
            logger.error(f"File Parse Error: {e}")
            raise

    def find_effort_header(self, df):
        """공수산정서 헤더 행 탐지: '순번', '항목', '공수' 키워드 기반"""
        for i, row in df.iterrows():
            row_str = ' '.join([str(v) for v in row if str(v) != 'nan'])
            if ('순번' in row_str or '번호' in row_str) and ('항목' in row_str or '내역' in row_str):
                return i
        return 0

    def find_effort_col(self, columns):
        """공수(M/D) 컬럼 자동 탐지"""
        for c in columns:
            cs = str(c)
            if '공수' in cs or 'M/D' in cs or 'MD' in cs or 'man' in cs.lower():
                return c
        return None

    def find_item_col(self, columns):
        """항목명 컬럼 자동 탐지"""
        for c in columns:
            cs = str(c)
            if '항목' in cs or '내역' in cs or '설명' in cs:
                return c
        return columns[1] if len(columns) > 1 else columns[0]

    def process_effort_to_quote(self, file_path, unit_price, template_path=None):
        """공수산정내역서 → 가온아이 표준 견적서 템플릿 기반 동적 주입"""
        if not template_path:
            template_path = PROJECT_ROOT / "template" / "Standard_Quote_Template.xlsx"
            
        output_path = QUOTE_OUTPUT_DIR / f"견적초안_{file_path.stem}.xlsx"
        
        try:
            # 1. 공수산정서 데이터 파싱
            df_raw = pd.read_excel(file_path, header=None)
            header_idx = self.find_effort_header(df_raw)
            effort_df = pd.read_excel(file_path, header=header_idx)

            item_col = self.find_item_col(effort_df.columns)
            effort_col = self.find_effort_col(effort_df.columns)

            items = []
            for _, row in effort_df.iterrows():
                item_name = row[item_col] if item_col else None
                md_val = row[effort_col] if effort_col else None

                if pd.isna(item_name) or str(item_name).strip() == '':
                    continue
                if pd.isna(md_val):
                    continue
                try:
                    md_float = float(md_val)
                except (ValueError, TypeError):
                    continue
                if md_float <= 0:
                    continue

                items.append({
                    'name': str(item_name).strip(),
                    'md': md_float,
                    'amount': md_float * unit_price
                })

            if not items:
                raise ValueError("공수산정서에서 유효한 공수 데이터를 추출하지 못했습니다.")

            # 2. 템플릿 파일 복사 및 로드
            shutil.copy(template_path, output_path)
            
            # 스타일 편집을 위해 openpyxl로 로드
            wb = load_workbook(output_path)
            ws = wb.active

            # 3. 고객사명 및 제안건명 추출 (NFC 한글 자모 분리 정규화 적용)
            import unicodedata
            filename_clean = unicodedata.normalize('NFC', file_path.stem)
            customer_name = "유한양행"  # 기본값
            if "유한양행" in filename_clean:
                customer_name = "유한양행"
            elif "NH투자증권" in filename_clean or "NH" in filename_clean:
                customer_name = "NH투자증권"
            else:
                parts = filename_clean.replace('_', ' ').replace('-', ' ').split()
                if parts:
                    customer_name = parts[0]

            # 제안건명 추출
            proposal_subject = "그룹웨어시스템 기능개발" # 기본값
            if "유한양행" in filename_clean:
                proposal_subject = "신규양식 2종 포함 총 양식 6종 및 SSO 모듈 추가 기술지원"
            else:
                proposal_subject = "그룹웨어 추가개발 기술지원"

            customer_name = unicodedata.normalize('NFC', customer_name)
            proposal_subject = unicodedata.normalize('NFC', proposal_subject)

            # 4. 상단 메타데이터 주입
            # D4 셀: 수신처
            ws['D4'] = customer_name
            
            # B9 셀: 제안서 서두 문구
            ws['B9'] = f'{customer_name} "{proposal_subject}"건으로 아래와 같이 견적을 제출 합니다. '
            
            # 5. 세부 품목 주입 및 행 확장 처리
            start_row = 14
            default_rows = 4  # 템플릿의 기본 내역 행 수 (14~17행)
            num_items = len(items)

            from copy import copy
            def copy_cell_style(src, dst):
                if src.has_style:
                    dst.font = copy(src.font)
                    dst.border = copy(src.border)
                    dst.fill = copy(src.fill)
                    dst.number_format = copy(src.number_format)
                    dst.protection = copy(src.protection)
                    dst.alignment = copy(src.alignment)

            # B열 병합 영역 해제 (구분 '개발기술지원' 병합 해제하여 행 삽입 시 레이아웃 깨짐 방지)
            target_range_b = None
            for rng in list(ws.merged_cells.ranges):
                if rng.bounds[0] == 2 and rng.bounds[1] == 13:  # min_col=2, min_row=13 (B13)
                    target_range_b = rng
                    break
            if target_range_b:
                ws.unmerge_cells(str(target_range_b))

            # G열 병합 영역 해제 (소비자가 단가 병합 해제하여 행 삽입 시 레이아웃 깨짐 방지)
            target_range_g = None
            for rng in list(ws.merged_cells.ranges):
                if rng.bounds[0] == 7 and rng.bounds[1] == 14:  # min_col=7, min_row=14 (G14)
                    target_range_g = rng
                    break
            if target_range_g:
                ws.unmerge_cells(str(target_range_g))

            # I열 병합 영역 해제 (제안금액 병합 해제하여 행 삽입 시 레이아웃 깨짐 방지)
            target_range_i = None
            for rng in list(ws.merged_cells.ranges):
                if rng.bounds[0] == 9 and rng.bounds[1] == 14:  # min_col=9, min_row=14 (I14)
                    target_range_i = rng
                    break
            if target_range_i:
                ws.unmerge_cells(str(target_range_i))

            # 품목이 4개보다 많으면 15행 자리에 동적으로 행 삽입
            if num_items > default_rows:
                insert_count = num_items - default_rows
                ws.insert_rows(start_row + 1, amount=insert_count)
                # 삽입한 행들에 스타일 복사 (14행 기준)
                for r_offset in range(1, insert_count + 1):
                    target_row = start_row + r_offset
                    for col in range(1, 12):
                        src_cell = ws.cell(row=start_row, column=col)
                        dst_cell = ws.cell(row=target_row, column=col)
                        copy_cell_style(src_cell, dst_cell)

            # 데이터 기입
            total_amount = 0
            for idx, item in enumerate(items):
                r = start_row + idx
                
                # B열: 구분 (템플릿 해제 후 일괄 재병합되므로 명시적 값은 B13에 존재)
                if idx == 0:
                    ws.cell(row=13, column=2, value='개발기술지원')

                # C열: 기능별
                ws.cell(row=r, column=3, value=None)
                
                # D열: 내역 (상세 품목명)
                ws.cell(row=r, column=4, value=f'* {item["name"]}')
                
                # F열: M/M (수량) [1M/M = 20M/D]
                ws.cell(row=r, column=6, value=item['md'] / 20.0)
                
                # G열: 단가 (1M/M 단가 = M/D 단가 * 20)
                if r == 14:
                    ws.cell(row=14, column=7, value=unit_price * 20)
                    ws.cell(row=14, column=7).number_format = '#,##0'
                
                # H열: 합계 (수식 = F * G14) -> G열이 병합되므로 모든 행이 G14 셀을 곱하도록 지정!
                ws.cell(row=r, column=8, value=f'=F{r}*G14')
                ws.cell(row=r, column=8).number_format = '#,##0'
                
                # I열: 제안금액 (I14에만 합계 수식을 넣고 나중에 일괄 재병합하므로 r=14 일 때만 기입)
                if r == 14:
                    ws.cell(row=14, column=9, value=f'=SUM(H14:H{start_row + num_items - 1})')
                    ws.cell(row=14, column=9).number_format = '#,##0'
                
                total_amount += item['amount']

            # 만약 품목이 4개보다 적으면 남는 템플릿 행은 빈 값으로 초기화 (기능별 C열 및 항목구분 D열은 레이아웃 유지 위해 보존)
            if num_items < default_rows:
                for r in range(start_row + num_items, start_row + default_rows):
                    ws.cell(row=r, column=6, value=None)
                    ws.cell(row=r, column=8, value=None)

            # 6. 소계 행들 탐색 및 업데이트 (위치가 변동되므로 동적 검색)
            soke_row = None
            grand_soke_row = None
            proposal_total_row = None

            max_search_row = 40 + max(0, num_items - default_rows)
            for r in range(start_row + num_items, max_search_row):
                c_val = ws.cell(row=r, column=3).value
                b_val = ws.cell(row=r, column=2).value
                
                # '가온아이 노임단가 기준 소계' 행 감지
                if c_val and '가온아이 노임단가 기준 소계' in str(c_val):
                    soke_row = r
                # '소 계' 행 감지
                elif c_val and '소 계' in str(c_val):
                    grand_soke_row = r
                # '제  안  금  액  합  계 (V.A.T 별도)' 행 감지
                elif b_val and '제  안  금  액  합  계' in str(b_val):
                    proposal_total_row = r

            # 소계 행 데이터 주입
            if soke_row:
                ws.cell(row=soke_row, column=6, value=f'=SUM(F14:F{start_row + num_items - 1})')
                ws.cell(row=soke_row, column=8, value=f'=SUM(H14:H{start_row + num_items - 1})')
                ws.cell(row=soke_row, column=8).number_format = '#,##0'
                
                # B열 구분란('개발기술지원') 동적 재병합
                ws.merge_cells(start_row=13, start_column=2, end_row=soke_row - 1, end_column=2)
                
                # G열 단가란(소비자가 단가) 동적 재병합
                ws.merge_cells(start_row=14, start_column=7, end_row=soke_row - 1, end_column=7)

            if grand_soke_row:
                ws.cell(row=grand_soke_row, column=9, value=f'=SUM(I14:I{start_row + num_items - 1})')
                ws.cell(row=grand_soke_row, column=9).number_format = '#,##0'

            if proposal_total_row:
                if grand_soke_row:
                    ws.cell(row=proposal_total_row, column=9, value=f'=I{grand_soke_row}')
                else:
                    ws.cell(row=proposal_total_row, column=9, value=total_amount)
                ws.cell(row=proposal_total_row, column=9).number_format = '#,##0'

            # 7. 제안금액 한글/숫자 표기 주입 (B10 셀)
            def num_to_kor(num):
                units = ["", "십", "백", "천"]
                big_units = ["", "만", "억", "조"]
                digits = ["", "일", "이", "삼", "사", "오", "육", "칠", "팔", "구"]
                
                if num == 0:
                    return "영"
                
                num_str = str(int(num))
                result = ""
                
                chunks = []
                while num_str:
                    chunks.append(num_str[-4:])
                    num_str = num_str[:-4]
                    
                for chunk_idx, chunk in enumerate(chunks):
                    chunk_res = ""
                    for digit_idx, digit in enumerate(reversed(chunk)):
                        d = int(digit)
                        if d != 0:
                            chunk_res = digits[d] + units[digit_idx] + chunk_res
                    if chunk_res:
                        result = chunk_res + big_units[chunk_idx] + result
                return result

            amount_kor = num_to_kor(total_amount)
            ws['B10'] = f'제안금액 :일금{amount_kor}원정 (₩{total_amount:,.0f}) V.A.T 별도'

            # 8. 비고란 수신처 치환 (B27 부근)
            bigo_row = None
            for r in range(max(start_row + num_items, grand_soke_row or 0), max_search_row):
                b_val = ws.cell(row=r, column=2).value
                if b_val and '비' in str(b_val) and '고' in str(b_val):
                    bigo_row = r
                    break

            if bigo_row:
                bigo_cell = ws.cell(row=bigo_row, column=3)
                if bigo_cell.value:
                    bigo_text = unicodedata.normalize('NFC', str(bigo_cell.value))
                    bigo_cell.value = bigo_text.replace("유한양행", customer_name)

            # 9. 최종 저장 및 가공 완료

            wb.save(output_path)
            logger.info(f"가온아이 표준 견적서 초안 생성 성공: {output_path} ({num_items}건, 합계: ₩{total_amount:,.0f})")
            return str(output_path)
        except Exception as e:
            logger.error(f"Effort→Quote Error: {e}")
            raise

manager = AutoPOManager()

@app.route('/')
def index():
    return send_file(WEB_DIR / 'index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
        
    file_path = INPUT_DIR / file.filename
    file.save(file_path)
    
    try:
        out_path = manager.process_file(file_path)
        return send_file(out_path, as_attachment=True)
    except Exception as e:
        logger.error(f"Upload process failed: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/upload_quote', methods=['POST'])
def upload_quote():
    if 'file' not in request.files:
        return jsonify({"error": "No file"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    unit_price = request.form.get('unit_price', 1350000)
    try:
        unit_price = float(unit_price)
    except (ValueError, TypeError):
        unit_price = 1350000

    file_path = EFFORT_INPUT_DIR / file.filename
    file.save(file_path)

    try:
        out_path = manager.process_effort_to_quote(file_path, unit_price)
        return send_file(out_path, as_attachment=True)
    except Exception as e:
        logger.error(f"Quote generation failed: {str(e)}")
        return jsonify({"error": str(e)}), 500

if __name__ == '__main__':
    logger.info("🚀 서버 시작: http://localhost:5001")
    app.run(port=5001, debug=True)
